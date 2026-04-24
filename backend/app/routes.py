import os
import base64
import time
import io
import datetime
import json
import math
from openpyxl import load_workbook

from flask import Blueprint, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
from werkzeug.exceptions import RequestEntityTooLarge

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.crypto import aes_engine
from app.ga import ga_state
from app.ga.ga_keygen import ga_generate_key_hex, TraceConfig
from app.utils.logger import app_logger

print(">>> ROUTES.PY YANG DIPAKAI:", __file__)

bp = Blueprint("main", __name__)

ALLOWED_EXTENSIONS = {"pdf", "docx"}

def allowed_file(filename: str) -> bool:
    if not filename or "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower()
    return ext in ALLOWED_EXTENSIONS

def bytes_to_bitstring(data: bytes) -> str:
    return "".join(f"{b:08b}" for b in data)

# =========================
# ERROR HANDLER
# =========================
@bp.errorhandler(RequestEntityTooLarge)
def handle_file_too_large(e):
    return jsonify(ok=False, error="Ukuran file melebihi 5MB (termasuk metadata)."), 413

# =========================
# PAGES
# =========================
@bp.get("/")
def index():
    return render_template("index.html")

@bp.get("/aes")
def aes():
    return render_template("aes.html")

# =========================
# AES BASELINE - GENERATE KEY
# =========================
@bp.post("/api/aes/generate_key")
def aes_generate_key():
    if "file" not in request.files:
        return jsonify(ok=False, error="File tidak ditemukan."), 400

    f = request.files["file"]
    if not f or f.filename == "":
        return jsonify(ok=False, error="Nama file kosong."), 400

    if not allowed_file(f.filename):
        return jsonify(ok=False, error="Format file harus .pdf atau .docx"), 400

    raw = f.read()
    if not raw:
        return jsonify(ok=False, error="File kosong."), 400

    try:
        # 🔑 KEY diturunkan dari seluruh isi file (deterministik)
        key_hex, key_b64 = aes_engine.derive_key_from_document(raw)

    except Exception as e:
        return jsonify(ok=False, error=f"Gagal generate key: {str(e)}"), 400

    return jsonify(
        ok=True,
        key_b64=key_b64,
        key_hex=key_hex,
        key_info={"bytes": 16, "bits": 128},
        filename=secure_filename(f.filename),
        size=len(raw)
    ), 200

# =========================
# AES BASELINE - ENCRYPT PREP (Visualization)
# =========================
@bp.post("/api/aes/encrypt_prep")
def aes_encrypt_prep():
    data = request.get_json(silent=True) or {}
    key_b64 = data.get("key_b64", "").strip()

    if not key_b64:
        return jsonify(ok=False, error="key_b64 tidak ditemukan"), 400

    try:
        key_bytes = base64.b64decode(key_b64)
        round_keys = aes_engine.aes_key_expansion(key_bytes)

        return jsonify(
            ok=True,
            round_keys=[
                {"round": i, "key_hex": rk} for i, rk in enumerate(round_keys)
            ],
            steps_info={
                "initial": {
                    "name": "Initial Round (Round 0)",
                    "operation": "AddRoundKey",
                    "description": "Plaintext dikombinasikan dengan Round Key 0 (kunci awal)."
                },
                "main": {
                    "name": "Main Rounds (Round 1-9)",
                    "operation": "SubBytes → ShiftRows → MixColumns → AddRoundKey",
                    "description": "Dilakukan sebanyak 9 kali untuk mentransformasi data secara bertahap."
                },
                "final": {
                    "name": "Final Round (Round 10)",
                    "operation": "SubBytes → ShiftRows → AddRoundKey",
                    "description": "Putaran terakhir tanpa operasi MixColumns."
                }
            }
        ), 200
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 400

# =========================
# AES BASELINE - ENCRYPT
# =========================
@bp.post("/api/aes/encrypt_file")
def aes_encrypt_file():
    if "file" not in request.files:
        return jsonify(ok=False, error="File tidak ditemukan."), 400

    f = request.files["file"]
    if not f or f.filename == "":
        return jsonify(ok=False, error="Nama file kosong."), 400

    if not allowed_file(f.filename):
        return jsonify(ok=False, error="Format file harus .pdf atau .docx"), 400

    raw = f.read()
    if not raw:
        return jsonify(ok=False, error="File kosong."), 400

    # 🔑 Gunakan kunci dari form jika ada, atau generate dari file
    key_b64_from_form = request.form.get("key_b64")

    try:
        t0 = time.perf_counter_ns()

        if key_b64_from_form:
            # Gunakan kunci yang sudah di-generate sebelumnya
            key_bytes = base64.b64decode(key_b64_from_form)
            res = aes_engine.encrypt_aes_baseline_bytes(raw, key_bytes)
        else:
            # Generate kunci dari file (legacy mode)
            res = aes_engine.encrypt_document_bytes_with_derived_key(raw)

        enc_ms = round((time.perf_counter_ns() - t0) / 1_000_000, 6)

    except Exception as e:
        return jsonify(ok=False, error=f"Gagal encrypt: {str(e)}"), 400

    package = {
    "mode": "AES-128-BASELINE",
    "filename": secure_filename(f.filename),
    "cipher_b64": res.cipher_b64,
    "key": res.key_b64,         # tetap dipertahankan agar decrypt lama tidak rusak
    "key_b64": res.key_b64,
    "key_hex": res.key_hex,
    "key_info": {
        "bytes": 16,
        "bits": 128
    },
    "meta": {
        "encrypt_ms": enc_ms,
        "padding": "PKCS7",
    }
        }

    # ✅ DIHAPUS: pemanggilan aes_export_xlsx() dari sini
    # Karena export_xlsx butuh JSON payload dari client.
    # History akan disimpan ketika user klik tombol "Save History (XLSX)".

    return jsonify(ok=True, package=package), 200

# =========================
# AES BASELINE - DECRYPT (JSON)
# =========================
@bp.post("/api/aes/decrypt_file")
def aes_decrypt_file():
    data = request.get_json(silent=True) or {}

    cipher_b64 = data.get("cipher_b64")
    key = data.get("key_b64") or data.get("key")
    orig_filename = data.get("filename") or "decrypted_file"

    if not all([cipher_b64, key]):
        return jsonify(ok=False, error="cipher/key tidak lengkap"), 400

    safe_name = secure_filename(orig_filename)
    if not safe_name:
        safe_name = "decrypted_file"

    # Tentukan tipe file berdasarkan nama file asli
    file_type = "PDF" if safe_name.lower().endswith(".pdf") else "DOCX"

    try:
        t0 = time.perf_counter_ns()

        # STEP 1: Decode Base64
        cipher_bytes = base64.b64decode(cipher_b64)

        # STEP 2: AES Decrypt
        key_bytes = base64.b64decode(key)
        from Crypto.Cipher import AES
        if len(cipher_bytes) < 32:
            raise ValueError("Ciphertext too short")
        iv = cipher_bytes[:16]
        ct = cipher_bytes[16:]
        cipher = AES.new(key_bytes, AES.MODE_CBC, iv)
        padded_plaintext = cipher.decrypt(ct)

        # STEP 3: Unpadding (PKCS7)
        pad_len = padded_plaintext[-1]
        if pad_len < 1 or pad_len > 16:
            raise ValueError("Invalid padding length")
        if padded_plaintext[-pad_len:] != bytes([pad_len]) * pad_len:
            raise ValueError("Invalid padding bytes")
        plaintext_bytes = padded_plaintext[:-pad_len]

        dec_ms = round((time.perf_counter_ns() - t0) / 1_000_000, 6)

        # STEP 4: Siapkan file untuk download
        bio = io.BytesIO(plaintext_bytes)
        bio.seek(0)

        # Tentukan MIME type
        if safe_name.lower().endswith(".pdf"):
            mimetype = "application/pdf"
        elif safe_name.lower().endswith(".docx"):
            mimetype = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        else:
            mimetype = "application/octet-stream"

    except Exception as e:
        app_logger.error(f"Dekripsi gagal: {str(e)}", exc_info=True)
        return jsonify(ok=False, error=f"Dekripsi gagal: {str(e)}"), 400

    # Kembalikan status step-by-step
    resp = send_file(
        bio,
        as_attachment=True,
        download_name=safe_name,
        mimetype=mimetype
    )

    # Tambahkan header dengan info dekripsi
    resp.headers["X-Dec-Ms"] = str(dec_ms)
    resp.headers["X-File-Type"] = file_type
    resp.headers["X-File-Size"] = str(len(plaintext_bytes))
    resp.headers["X-Padding-Status"] = "Removed (PKCS7)"
    resp.headers["X-Decrypt-Status"] = "Success"

    return resp

# =========================
# AES BASELINE - DECRYPT (JSON - untuk menampilkan step-by-step)
# =========================


@bp.post("/api/aes/decrypt_info")
def aes_decrypt_info():
    data = request.get_json(silent=True) or {}

    cipher_b64 = data.get("cipher_b64", "").strip()
    key_b64 = data.get("key_b64", "").strip()
    filename = data.get("filename", "decrypted_file")

    if not cipher_b64 or not key_b64:
        return jsonify(ok=False, error="cipher_b64/key_b64 tidak lengkap"), 400

    file_type = "PDF" if filename.lower().endswith(".pdf") else "DOCX"

    try:
        steps = []

        # STEP 1: Decode Base64
        steps.append({
            "step": 1,
            "name": "Decode Base64",
            "status": "Processing...",
            "details": f"Mendekode {len(cipher_b64)} karakter Base64 menjadi bytes"
        })
        cipher_bytes = base64.b64decode(cipher_b64)
        steps[-1]["status"] = "✅ Success"
        steps[-1]["details"] = f"Berhasil mendekode {len(cipher_bytes)} bytes ciphertext"

        # STEP 2: AES Decrypt
        steps.append({
            "step": 2,
            "name": "AES Decrypt (CBC Mode)",
            "status": "Processing...",
            "details": f"Decrypt dengan AES-128-CBC menggunakan kunci {len(base64.b64decode(key_b64))} byte"
        })
        key_bytes = base64.b64decode(key_b64)
        from Crypto.Cipher import AES
        if len(cipher_bytes) < 32:
            raise ValueError("Ciphertext too short (missing IV)")
        iv = cipher_bytes[:16]
        ct = cipher_bytes[16:]
        cipher = AES.new(key_bytes, AES.MODE_CBC, iv)
        padded_plaintext = cipher.decrypt(ct)
        steps[-1]["status"] = "✅ Success"
        steps[-1]["details"] = f"Berhasil decrypt: {len(padded_plaintext)} bytes (masih ter-padding)"

        # STEP 3: Unpadding (PKCS7)
        steps.append({
            "step": 3,
            "name": "Unpadding (PKCS7)",
            "status": "Processing...",
            "details": f"Menghapus padding PKCS7 ({padded_plaintext[-1]} bytes)"
        })
        pad_len = padded_plaintext[-1]
        if pad_len < 1 or pad_len > 16:
            raise ValueError("Invalid padding length")
        if padded_plaintext[-pad_len:] != bytes([pad_len]) * pad_len:
            raise ValueError("Invalid padding bytes")
        plaintext_bytes = padded_plaintext[:-pad_len]
        steps[-1]["status"] = "✅ Success"
        steps[-1]["details"] = f"Padding dihapus: {len(plaintext_bytes)} bytes plaintext asli"

        # STEP 4: File Info
        steps.append({
            "step": 4,
            "name": "Rekonstruksi File",
            "status": "✅ Success",
            "details": f"File {file_type} siap didownload ({len(plaintext_bytes)} bytes)"
        })

        file_size_kb = round(len(plaintext_bytes) / 1024, 2)

        # Generate key expansion (11 round keys)
        key_bytes = base64.b64decode(key_b64)
        round_keys = aes_engine.aes_key_expansion(key_bytes)

        return jsonify(
            ok=True,
            steps=steps,
            file_info={
                "type": file_type,
                "size_bytes": len(plaintext_bytes),
                "size_kb": file_size_kb,
                "filename": filename
            },
            cipher_preview=cipher_b64[:100] + "..." if len(cipher_b64) > 100 else cipher_b64,
            key_expansion={
                "original_key": key_bytes.hex().upper(),
                "round_keys": [
                    {"round": i, "key_hex": rk} for i, rk in enumerate(round_keys)
                ],
                "total_rounds": 10
            },
            round_summary={
                "rounds_1_to_9": {
                    "count": 9,
                    "operations": ["InvShiftRows", "InvSubBytes", "InvMixColumns", "AddRoundKey"],
                    "description": "Round 1-9: 4 operasi invers AES"
                }
            },
            final_round={
                "round": 10,
                "operations": ["InvShiftRows", "InvSubBytes", "AddRoundKey"],
                "note": "Tanpa InvMixColumns",
                "description": "Final Round: 3 operasi invers AES"
            },
            padding_info={
                "padding_type": "PKCS7",
                "padding_bytes": pad_len,
                "plaintext_bytes": len(plaintext_bytes)
            }
        ), 200

    except Exception as e:
        app_logger.error(f"Dekripsi gagal: {str(e)}", exc_info=True)
        return jsonify(ok=False, error=f"Dekripsi gagal: {str(e)}"), 400

# =========================
# AES BASELINE - DECRYPT PACKAGE
# =========================
@bp.post("/api/aes/decrypt_package")
def aes_decrypt_package():
    data = request.get_json(silent=True) or {}
    pkg = data.get("package")

    if not pkg or not isinstance(pkg, dict):
        return jsonify(ok=False, error="Package tidak ditemukan."), 400

    cipher_b64 = pkg.get("cipher_b64")
    key = pkg.get("key_b64") or pkg.get("key")
    orig_filename = (pkg.get("filename") or "").strip()

    if not all([cipher_b64, key]):
        return jsonify(ok=False, error="Isi package kurang (cipher/key)."), 400

    safe_name = secure_filename(orig_filename) if orig_filename else ""
    if not safe_name:
        safe_name = "decrypted_file"

    mimetype = "application/octet-stream"
    lower = safe_name.lower()
    if lower.endswith(".pdf"):
        mimetype = "application/pdf"
    elif lower.endswith(".docx"):
        mimetype = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"

    try:
        t0 = time.perf_counter_ns()
        pt_bytes = aes_engine.decrypt_aes_baseline_bytes(cipher_b64, key)
        dec_ms = round((time.perf_counter_ns() - t0) / 1_000_000, 6)
    except Exception:
        return jsonify(ok=False, error="Dekripsi gagal: paket/kunci tidak valid atau data rusak."), 400

    bio = io.BytesIO(pt_bytes)
    bio.seek(0)

    resp = send_file(
        bio,
        as_attachment=True,
        download_name=safe_name,
        mimetype=mimetype
    )
    resp.headers["X-Dec-Ms"] = str(dec_ms)
    return resp

# =========================
# AES EXPORT XLSX
# =========================
@bp.post("/api/aes/export_xlsx")
def aes_export_xlsx():
    data = request.get_json(silent=True)
    if not data:
        return jsonify(ok=False, error="Data tidak valid"), 400

    # key_hex
    try:
        key_hex = (data.get("key_hex") or "").strip()
        if not key_hex:
            key_hex = base64.b64decode(data.get("key", "")).hex()
    except Exception:
        key_hex = "-"

    file_path = "aes_baseline_history.xlsx"

    headers = [
        "timestamp",
        "mode",
        "file",
        "ukuran_kb",
        "key_hex",
        "encrypt_ms",
        "decrypt_ms",
        "entropy_cipher",
        "nist_frequency",
        "nist_runs",
        "nist_freq_p",
        "nist_runs_p",
        "catatan"
    ]

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        # jika sheet kosong, tulis header
        if ws.max_row == 0:
            ws.append(headers)
        # kalau header lama tidak sama, tetap aman: kalau baris 1 bukan header, tambahkan header baru
        else:
            first_row = [c.value for c in ws[1]]
            if first_row != headers:
                # jangan merusak file lama, tambahkan header versi baru di baris berikutnya sebagai penanda
                # (opsional) kamu bisa hapus ini kalau ingin strict
                pass
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "AES_History"
        ws.append(headers)

    ws.append([
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        data.get("mode", "AES-128-BASELINE"),
        data.get("filename", "-"),
        data.get("size_kb", "-"),
        key_hex,
        data.get("encrypt_ms", "-"),
        data.get("decrypt_ms", "-"),
        data.get("entropy_cipher", "-"),
        data.get("nist_frequency", "-"),
        data.get("nist_runs", "-"),
        data.get("nist_freq_p", "-"),
        data.get("nist_runs_p", "-"),
        ""
    ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    wb.save(file_path)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="aes_baseline_history.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

#==========================#
# AES TEST
#==========================#

def _b64_to_bytes(b64str: str) -> bytes:
    try:
        return base64.b64decode(b64str, validate=True)
    except Exception:
        # fallback kalau ada whitespace/newline
        return base64.b64decode(b64str)

def shannon_entropy_bytes(data: bytes) -> float:
    if not data:
        return 0.0
    freq = [0] * 256
    for b in data:
        freq[b] += 1
    n = len(data)
    ent = 0.0
    for c in freq:
        if c == 0:
            continue
        p = c / n
        ent -= p * math.log2(p)
    return ent  # dalam bit per byte (maks ~8)

def bytes_to_bitstring(data: bytes) -> str:
    return "".join(f"{b:08b}" for b in data)

def nist_frequency_monobit(bits: str) -> dict:
    # NIST SP800-22 Frequency (Monobit) Test
    n = len(bits)
    if n == 0:
        return {"pass": False, "p_value": 0.0, "reason": "empty bits"}
    s = 0
    for ch in bits:
        s += 1 if ch == "1" else -1
    sobs = abs(s) / math.sqrt(n)
    p_value = math.erfc(sobs / math.sqrt(2))
    return {"pass": p_value >= 0.01, "p_value": p_value}


def nist_runs_test(bits: str) -> dict:
    # NIST SP800-22 Runs Test (requires frequency test precondition)
    n = len(bits)
    if n < 2:
        return {"pass": False, "p_value": 0.0, "reason": "bits too short"}

    ones = bits.count("1")
    pi = ones / n

    # precondition: |pi - 0.5| < 2/sqrt(n)
    tau = 2.0 / math.sqrt(n)
    if abs(pi - 0.5) >= tau:
        return {"pass": False, "p_value": 0.0, "reason": "pi too far from 0.5"}

    # count runs
    v = 1
    for i in range(1, n):
        if bits[i] != bits[i - 1]:
            v += 1

    num = abs(v - 2.0 * n * pi * (1.0 - pi))
    den = 2.0 * math.sqrt(2.0 * n) * pi * (1.0 - pi)
    p_value = math.erfc(num / den)
    return {"pass": p_value >= 0.01, "p_value": p_value}

@bp.post("/api/aes/test_entropy")
def aes_test_entropy():
    data = request.get_json(silent=True) or {}
    cipher_b64 = (data.get("cipher_b64") or "").strip()
    if not cipher_b64:
        return jsonify(ok=False, error="cipher_b64 tidak ada"), 400

    try:
        ct = _b64_to_bytes(cipher_b64)
        ent = shannon_entropy_bytes(ct)
    except Exception as e:
        return jsonify(ok=False, error=f"Gagal hitung entropy: {str(e)}"), 400

    return jsonify(ok=True, entropy=ent), 200

@bp.post("/api/aes/test_nist")
def aes_test_nist():
    data = request.get_json(silent=True) or {}
    cipher_b64 = (data.get("cipher_b64") or "").strip()
    if not cipher_b64:
        return jsonify(ok=False, error="cipher_b64 tidak ada"), 400

    try:
        ct = _b64_to_bytes(cipher_b64)
        bits = bytes_to_bitstring(ct)

        # NIST butuh panjang bit memadai; kalau terlalu kecil hasilnya kurang meaningful
        if len(bits) < 1000:
            # tetap boleh, tapi beri warning di response
            warning = f"Panjang bit {len(bits)} < 1000, hasil uji kurang stabil."
        else:
            warning = ""

        freq = nist_frequency_monobit(bits)
        runs = nist_runs_test(bits)

    except Exception as e:
        return jsonify(ok=False, error=f"Gagal NIST test: {str(e)}"), 400

    return jsonify(ok=True, frequency=freq, runs=runs, warning=warning), 200

@bp.post("/api/aesga/generate_key")
def aesga_generate_key():
    if "file" not in request.files:
        return jsonify(ok=False, error="File tidak ditemukan."), 400

    f = request.files["file"]
    if not f or f.filename == "":
        return jsonify(ok=False, error="Nama file kosong."), 400

    if not allowed_file(f.filename):
        return jsonify(ok=False, error="Format file harus .pdf atau .docx"), 400

    raw = f.read()
    if not raw:
        return jsonify(ok=False, error="File kosong."), 400

    # sesuai metodologi proposal: konversi biner (minimal raw->bitstring)
    plaintext = ""  # opsional dikembangkan: ekstraksi teks PDF/DOCX
    bitstring = bytes_to_bitstring(raw)

    sid = ga_state.new_session(
        filename=secure_filename(f.filename),
        mime=f.mimetype or "application/octet-stream",
        size=len(raw),
        plaintext=plaintext,
        bitstring=bitstring,
    )

    # GA params (Optimized defaults for Speed & Quality)
    pop = int(request.form.get("population", 60))
    gens = int(request.form.get("generations", 50))
    cr = float(request.form.get("crossover_rate", 0.8))
    mr = float(request.form.get("mutation_rate", 0.05))
    ft = float(request.form.get("fitness_threshold", 0.95))
    seed_val = request.form.get("seed")
    seed = int(seed_val) if seed_val not in (None, "", "null") else None

    tcfg = TraceConfig(
        enabled=True,
        top_n=int(request.form.get("trace_top_n", 5)),
        selection_pairs=int(request.form.get("trace_selection_pairs", 5)),
        crossover_events=int(request.form.get("trace_crossover_events", 5)),
        mutation_events=int(request.form.get("trace_mutation_events", 5)),
        key_preview_hex_chars=int(request.form.get("trace_preview_chars", 32)),
        )

    # Ekstrak tiga sampel (0%, 50%, 100%) agar cepat namun tetap robust
    sample_len = 2048
    plaintext_samples = []
    file_size = len(raw)
    num_samples = 3
    
    for i in range(num_samples):
        # Hitung posisi start secara merata
        start = (i * (file_size - sample_len)) // (num_samples - 1) if file_size > sample_len else 0
        start = (start // 16) * 16  # Align dengan blok AES 16-byte
        s = raw[start : start + sample_len]
        plaintext_samples.append(s.ljust(sample_len, b'\0') if len(s) < sample_len else s)

    try:
        # AES-128 key
        key_hex, meta, trace = ga_generate_key_hex(
            key_bits=128,
            population=pop,
            generations=gens,
            crossover_rate=cr,
            mutation_rate=mr,
            fitness_threshold=ft,
            seed=seed,
            trace_cfg=tcfg,
            plaintext_samples=plaintext_samples
        )

    except Exception as e:

        ga_state.delete_session(sid)
        return jsonify(ok=False, error=f"Gagal generate key GA: {str(e)}"), 400

    st = ga_state.get_session(sid)
    st.best_key_hex = key_hex
    st.ga_meta = meta
    st.history = trace

    key_bytes = bytes.fromhex(key_hex)
    key_b64 = base64.b64encode(key_bytes).decode("ascii")

    return jsonify(
        ok=True,
        session_id=sid,
        key_b64=key_b64,
        key_hex=key_hex,
        ga_meta=meta,
        trace=trace,
        bit_len=len(bitstring),
        size=len(raw),
    ), 200

# =========================
# AES+GA - ENCRYPT PREP (Visualization)
# =========================
@bp.post("/api/aesga/encrypt_prep")
def aesga_encrypt_prep():
    """
    Endpoint untuk menyiapkan data visualisasi enkripsi AES+GA.
    """
    data = request.get_json(silent=True) or {}
    key_b64 = data.get("key_b64", "").strip()

    if not key_b64:
        return jsonify(ok=False, error="key_b64 tidak ditemukan"), 400

    try:
        key_bytes = base64.b64decode(key_b64)
        round_keys = aes_engine.aes_key_expansion(key_bytes)

        return jsonify(
            ok=True,
            round_keys=[
                {"round": i, "key_hex": rk} for i, rk in enumerate(round_keys)
            ],
            steps_info={
                "initial": {
                    "name": "Initial Round (Round 0)",
                    "operation": "AddRoundKey",
                    "description": "Plaintext dikombinasikan dengan Round Key 0 (hasil optimasi GA)."
                },
                "main": {
                    "name": "Main Rounds (Round 1-9)",
                    "operation": "SubBytes → ShiftRows → MixColumns → AddRoundKey",
                    "description": "Dilakukan sebanyak 9 kali untuk mentransformasi data secara bertahap menggunakan kunci yang dioptimasi."
                },
                "final": {
                    "name": "Final Round (Round 10)",
                    "operation": "SubBytes → ShiftRows → AddRoundKey",
                    "description": "Putaran terakhir tanpa MixColumns, menghasilkan ciphertext final."
                }
            }
        ), 200
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 400

# =========================
# enkripsi file dengan AES+GA
# =========================

@bp.post("/api/aesga/encrypt_file")
def aesga_encrypt_file():
    if "file" not in request.files:
        return jsonify(ok=False, error="File tidak ditemukan."), 400

    f = request.files["file"]
    if not f or f.filename == "":
        return jsonify(ok=False, error="Nama file kosong."), 400

    if not allowed_file(f.filename):
        return jsonify(ok=False, error="Format file harus .pdf atau .docx"), 400

    key_b64 = (request.form.get("key_b64") or "").strip()
    if not key_b64:
        return jsonify(ok=False, error="key_b64 tidak ada. Generate key GA dulu."), 400

    raw = f.read()
    if not raw:
        return jsonify(ok=False, error="File kosong."), 400

    try:
        # decode key (harus 16 byte)
        key_bytes = base64.b64decode(key_b64)
        t0 = time.perf_counter_ns()

        # AES sama seperti baseline: ECB + PKCS7
        res = aes_engine.encrypt_aes_baseline_bytes(raw, key_bytes)

        enc_ms = round((time.perf_counter_ns() - t0) / 1_000_000, 6)
    except Exception as e:
        return jsonify(ok=False, error=f"Gagal encrypt AES+GA: {str(e)}"), 400

    package = {
    "mode": "AES-128-GA",
    "filename": secure_filename(f.filename),
    "cipher_b64": res.cipher_b64,
    "key": res.key_b64,
    "key_b64": res.key_b64,
    "key_hex": res.key_hex,
    "key_info": {
        "bytes": 16,
        "bits": 128
    },
    "meta": {
        "encrypt_ms": enc_ms,
        "padding": "PKCS7",
        "aes_mode": "CBC"
    }
}

    return jsonify(ok=True, package=package), 200

# =========================
# AES+GA - DECRYPT (JSON)  ✅ ADD THIS
# =========================
# AES+GA - DECRYPT FILE
# =========================
@bp.post("/api/aesga/decrypt_file")
def aesga_decrypt_file():
    data = request.get_json(silent=True) or {}

    cipher_b64 = (data.get("cipher_b64") or "").strip()
    key_b64 = (data.get("key_b64") or data.get("key") or "").strip()
    orig_filename = data.get("filename") or "decrypted_file"

    if not cipher_b64 or not key_b64:
        return jsonify(ok=False, error="cipher_b64/key_b64 tidak lengkap"), 400

    safe_name = secure_filename(orig_filename) or "decrypted_file"

    # Tentukan tipe file berdasarkan nama file asli
    file_type = "PDF" if safe_name.lower().endswith(".pdf") else "DOCX"

    try:
        t0 = time.perf_counter_ns()

        # STEP 1: Decode Base64 (ciphertext)
        cipher_bytes = base64.b64decode(cipher_b64)

        # STEP 2: Decode Base64 (key)
        key_bytes = base64.b64decode(key_b64)

        # STEP 3: AES Decrypt (CBC Mode)
        from Crypto.Cipher import AES
        if len(cipher_bytes) < 32:
            raise ValueError("Ciphertext too short (missing IV)")
        iv = cipher_bytes[:16]
        ct = cipher_bytes[16:]
        cipher = AES.new(key_bytes, AES.MODE_CBC, iv)
        padded_plaintext = cipher.decrypt(ct)

        # STEP 4: Unpadding (PKCS7) - Explicit validation
        pad_len = padded_plaintext[-1]
        if pad_len < 1 or pad_len > 16:
            raise ValueError("Invalid padding length")
        if padded_plaintext[-pad_len:] != bytes([pad_len]) * pad_len:
            raise ValueError("Invalid padding bytes")
        plaintext_bytes = padded_plaintext[:-pad_len]

        dec_ms = round((time.perf_counter_ns() - t0) / 1_000_000, 6)

        # STEP 5: Siapkan file untuk download
        bio = io.BytesIO(plaintext_bytes)
        bio.seek(0)

        # Tentukan MIME type
        if safe_name.lower().endswith(".pdf"):
            mimetype = "application/pdf"
        elif safe_name.lower().endswith(".docx"):
            mimetype = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        else:
            mimetype = "application/octet-stream"

    except Exception as e:
        app_logger.error(f"Dekripsi gagal: {str(e)}", exc_info=True)
        return jsonify(ok=False, error=f"Dekripsi gagal: {str(e)}"), 400

    # Kembalikan file dengan headers lengkap (seperti AES Baseline)
    resp = send_file(
        bio,
        as_attachment=True,
        download_name=safe_name,
        mimetype=mimetype
    )

    # Headers lengkap untuk monitoring & debugging
    resp.headers["X-Dec-Ms"] = str(dec_ms)
    resp.headers["X-File-Type"] = file_type
    resp.headers["X-File-Size"] = str(len(plaintext_bytes))
    resp.headers["X-Padding-Status"] = "Removed (PKCS7)"
    resp.headers["X-Decrypt-Status"] = "Success"

    return resp

# =========================
# AES+GA - DECRYPT INFO (Step-by-Step Visualization)
# =========================
@bp.post("/api/aesga/decrypt_info")
def aesga_decrypt_info():
    """
    Endpoint untuk menampilkan step-by-step proses dekripsi AES+GA dengan SweetAlert.
    Mirip seperti /api/aes/decrypt_info tapi untuk AES+GA.
    """
    data = request.get_json(silent=True) or {}

    cipher_b64 = (data.get("cipher_b64") or "").strip()
    key_b64 = (data.get("key_b64") or data.get("key") or "").strip()
    filename = data.get("filename", "decrypted_file")

    if not cipher_b64 or not key_b64:
        return jsonify(ok=False, error="cipher_b64/key_b64 tidak lengkap"), 400

    safe_name = secure_filename(filename) or "decrypted_file"
    file_type = "PDF" if safe_name.lower().endswith(".pdf") else "DOCX"

    try:
        steps = []

        # STEP 1: Decode Base64 (Ciphertext & Key)
        steps.append({
            "step": 1,
            "name": "Decode Base64",
            "status": "Processing...",
            "details": f"Mendekode ciphertext ({len(cipher_b64)} char) dan key menjadi bytes"
        })
        cipher_bytes = base64.b64decode(cipher_b64)
        key_bytes = base64.b64decode(key_b64)
        if len(key_bytes) != 16:
            raise ValueError(f"Key harus 16 byte, didapatkan {len(key_bytes)} byte")
        steps[-1]["status"] = "✅ Success"
        steps[-1]["details"] = f"Berhasil mendekode data: {len(cipher_bytes)} bytes ciphertext, {len(key_bytes)} bytes key"

        # STEP 2: AES Decrypt (CBC Mode)
        steps.append({
            "step": 2,
            "name": "AES Decrypt (CBC Mode)",
            "status": "Processing...",
            "details": f"Decrypt dengan AES-128-CBC menggunakan kunci {len(key_bytes)} byte"
        })
        from Crypto.Cipher import AES
        if len(cipher_bytes) < 32:
            raise ValueError("Ciphertext too short (missing IV)")
        iv = cipher_bytes[:16]
        ct = cipher_bytes[16:]
        cipher = AES.new(key_bytes, AES.MODE_CBC, iv)
        padded_plaintext = cipher.decrypt(ct)
        steps[-1]["status"] = "✅ Success"
        steps[-1]["details"] = f"Berhasil decrypt: {len(padded_plaintext)} bytes (ter-padding)"

        # STEP 3: Unpadding (PKCS7)
        steps.append({
            "step": 3,
            "name": "Unpadding (PKCS7)",
            "status": "Processing...",
            "details": f"Menghapus padding PKCS7 ({padded_plaintext[-1]} bytes)"
        })
        pad_len = padded_plaintext[-1]
        if pad_len < 1 or pad_len > 16:
            raise ValueError("Invalid padding length")
        if padded_plaintext[-pad_len:] != bytes([pad_len]) * pad_len:
            raise ValueError("Invalid padding bytes")
        plaintext_bytes = padded_plaintext[:-pad_len]
        steps[-1]["status"] = "✅ Success"
        steps[-1]["details"] = f"Padding dihapus: {len(plaintext_bytes)} bytes plaintext asli"

        # STEP 4: Rekonstruksi File
        steps.append({
            "step": 4,
            "name": "Rekonstruksi File",
            "status": "✅ Success",
            "details": f"File {file_type} siap didownload ({len(plaintext_bytes)} bytes)"
        })

        file_size_kb = round(len(plaintext_bytes) / 1024, 2)

        # Generate key expansion (11 round keys)
        round_keys = aes_engine.aes_key_expansion(key_bytes)

        return jsonify(
            ok=True,
            steps=steps,
            file_info={
                "type": file_type,
                "size_bytes": len(plaintext_bytes),
                "size_kb": file_size_kb,
                "filename": safe_name
            },
            cipher_preview=cipher_b64[:100] + "..." if len(cipher_b64) > 100 else cipher_b64,
            key_expansion={
                "original_key": key_bytes.hex().upper(),
                "round_keys": [
                    {"round": i, "key_hex": rk} for i, rk in enumerate(round_keys)
                ],
                "total_rounds": 10
            },
            round_summary={
                "rounds_1_to_9": {
                    "count": 9,
                    "operations": ["InvShiftRows", "InvSubBytes", "InvMixColumns", "AddRoundKey"],
                    "description": "Round 1-9: 4 operasi invers AES"
                }
            },
            final_round={
                "round": 10,
                "operations": ["InvShiftRows", "InvSubBytes", "AddRoundKey"],
                "note": "Tanpa InvMixColumns",
                "description": "Final Round: 3 operasi invers AES"
            },
            padding_info={
                "padding_type": "PKCS7",
                "padding_bytes": pad_len,
                "plaintext_bytes": len(plaintext_bytes)
            }
        ), 200

    except Exception as e:
        app_logger.error(f"Dekripsi gagal: {str(e)}", exc_info=True)
        return jsonify(ok=False, error=f"Dekripsi gagal: {str(e)}"), 400

#==========================
#return main menu AES+GA
#==========================

@bp.get("/aes_ga")
def aes_ga():
    return render_template("aes_ga.html")
